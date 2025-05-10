import json
import re
from collections import defaultdict
import nltk
from nltk.translate.meteor_score import meteor_score
from nltk.translate.bleu_score import corpus_bleu
from rdkit import Chem
import selfies as sf
import os
import numpy as np
import openpyxl
from tqdm.auto import tqdm
from smiles_metrics import calculate_smiles_metrics

def answer_sort(task_accuracy):
    
    return task_accuracy


def read_file(file):
    datas = {
        "客观": [],
        "主观": [],
        "分子翻译": [],
        "分子设计": [],
        "信息抽取": [],
        "工艺图": [],
        "装置图":[]
    }
    with open(file,'r',encoding='utf-8') as f:
        for line in f:
            item = json.loads(line)

            item["task"] = item["filename"]
            if "选择" in item["task"] or "判断" in item["task"] or "property_prediction" in item["task"] or "reaction_prediction" in item["task"] or "retro_synthesis" in item["task"] or "yield_prediction" in item["task"] or "MultipleChoice" in item["task"] or "TrueFalse" in item["task"]:
                datas["客观"].append(item)
            elif "填空" in item["task"] or "计算" in item["task"] or "简答" in item["task"] or "molecule_description" in item["task"] or "FillBlank" in item["task"] or "ShortAnswer" in item["task"] or "CalculationQuestions" in item["task"] or "高考" in item["task"] or "中考" in item["task"]:
                datas["主观"].append(item)
            elif "chem-smiles-to-iupac-cllm" in item["task"]:
                datas["分子翻译"].append(item)
            elif "molecule_design" in item["task"]:
                datas["分子设计"].append(item)
            elif "abstract_question" in item["task"]:
                datas["信息抽取"].append(item)
    return datas


def extract_final_score(answer):
    match = re.findall(r'\d+\.?\d*', answer)
    # match = re.findall(r'<最终得分>\s*([\d\.]+)', answer)
    if match:
        return float(match[-1])
    return None

def subjective_task(datas):

    task_scores = defaultdict(lambda: {'correct': 0, 'total': 0})
    i = 0
    for data in datas:
        answer = data.get('answer', '')
        task = data.get('task', '')
        # task = index

        final_score = extract_final_score(answer)
        if final_score is not None:
            task_scores[task]['total'] += 1

            # 填空
            if "FillBlank" in task or "填空" in task:
                if final_score > 1 or final_score < 0:
                    print(answer, 111111, i)
                task_scores[task]['correct'] += final_score

            # 简答计算
            else:
                if final_score > 5 or final_score < 0:
                    print(answer, i)

                task_scores[task]["correct"] += (final_score-1)/4
                # if final_score > 2:
                #     task_scores[task]['correct'] += 1
        i += 1

    # 特殊类别，需要单独处理
    lingyv_correct_f, lingyv_total_f, lingyv_correct_s, lingyv_total_s, lingyv_correct_c, lingyv_total_c = 0, 0, 0, 0, 0, 0
    # task_accuracy = {}
    for task, scores in task_scores.items():
        print(task, scores["total"])
        if "领域知识" in task and "填空" in task:
            lingyv_correct_f += scores['correct']
            lingyv_total_f += scores["total"]
        elif "领域知识" in task and "计算" in task:
            lingyv_correct_c += scores['correct']
            lingyv_total_c += scores["total"]
        elif "领域知识" in task and "简答" in task:
            lingyv_correct_s += scores['correct']
            lingyv_total_s += scores["total"]
        else:
            task_accuracy[task] = (scores['correct'] / scores['total']) * 100 
    if lingyv_correct_c != 0:
        print(lingyv_correct_c, lingyv_total_c)
        task_accuracy["领域知识_计算"] = (lingyv_correct_c / lingyv_total_c) * 100
    if lingyv_correct_s != 0:
        task_accuracy["领域知识_简答"] = (lingyv_correct_s / lingyv_total_s) * 100
    if lingyv_correct_f != 0:
        task_accuracy["领域知识_填空"] = (lingyv_correct_f / lingyv_total_f) * 100

    return task_accuracy
  
def objective_task(datas):
    task_scores = defaultdict(lambda: {'correct': 0, 'total': 0})
    for data in datas:
        task = data.get('task', '')
        task_accuracy = {}
        task_scores[task]['total'] += 1
        if "选择" in task or "MultipleChoice" in task:
            if predict in data.keys():
                data[predict] = data[predict].replace("<最终答案>", "").replace("\n", "").replace(",","").replace(" ","").replace("<最终选项>", "").replace("和", "").replace("、", "").replace("*", "").replace(";", "")
            else:
                data[predict] = None
            data[target] = data[target].replace("<最终答案>", "").replace("\n", "").replace(",","").replace(" ","").replace("<最终选项>", "").replace("和", "").replace("、", "").replace("*", "").replace(";", "")

        elif "判断" in task or "TrueFalse" in task:
            if ("正确" in data[predict] and "错误" in data[predict]) or ("正确" not in data[predict] and "错误" not in data[predict]):
                print(data[predict])
                data[predict] = None
            elif "正确" in data[predict]:
                data[predict] = "正确"
            elif "错误" in data[predict]:
                data[predict] = "错误"
            data[target] = data[target].replace("对", "正确").replace("错", "错误").replace("错误误", "错误").replace("Incorrect","错误").replace("Correct","正确")
            # print(data[predict], data[target])
            # data[predict] = data[predict]

        elif "property_prediction" in task or "reaction_prediction" in task or "retro_synthesis" in task or "yield_prediction" in task:
            if "<answer_start>正确<answer_end>" in data[predict]  and "<answer_start>错误<answer_end>" in data[predict]:
                data[predict] = "错误"
            elif "<answer_start>正确<answer_end>" in data[predict]:
                data[predict] = "正确"
            data[target] = "正确"
        
        if data[predict] == data[target]:
            # print(data[predict], data[target], task_scores[task]['correct'])
            task_scores[task]['correct'] += 1

    lingyv_correct_m, lingyv_total_m, lingyv_correct_t, lingyv_total_t = 0, 0, 0, 0
    for task, scores in task_scores.items():
        print(task, scores["total"])
        if "领域知识" in task and "选择" in task:
            lingyv_correct_m += scores['correct']
            lingyv_total_m += scores["total"]
        elif "领域知识" in task and "判断" in task:
            lingyv_correct_t += scores['correct']
            lingyv_total_t += scores["total"]
        else:
            task_accuracy[task] = (scores['correct'] / scores['total']) * 100

    if lingyv_correct_m != 0:
        task_accuracy["领域知识_选择"] = (lingyv_correct_m / lingyv_total_m) * 100
    if lingyv_correct_t != 0:
        task_accuracy["领域知识_判断"] = (lingyv_correct_t / lingyv_total_t) * 100
    return task_accuracy

def smiles2iupac(datas):
    references = []
    predictions = []
        # 遍历字典中的数据
    for i in range(len(datas)):
        gold = datas[i].get("gold", "")
        pre = datas[i].get("answer", "").replace("<抽取结果>\n","").replace("<抽取结果>","")
        predictions.append([pre])
        references.append([gold])

    # 检查参考答案和预测答案的数量是否相等
    assert len(references) == len(predictions), "参考答案和预测答案的数量必须相等"
    # 计算总的样本数
    total = len(references)
    # 初始化精确匹配的数量
    exact_match_count = 0

    # 遍历参考答案和预测答案
    for ref_list, pred_list in zip(references, predictions):

        # 将参考答案和预测答案都转换为小写 IUPAC不区分大小写
        ref_str = ref_list[0].lower()
        pred_str = pred_list[0].lower()
        # 统计分数
        if ref_str == pred_str:
            # 如果参考答案和预测答案相同，则将精确匹配的数量加1
            exact_match_count += 1

    # 计算精确匹配的分数
    exact_match_score = (exact_match_count / total)
    # 返回精确匹配的分数
    return {"分子翻译": 100 * exact_match_score}

def molecule_design(datas):
    golds = []
    pres = []
    # 遍历文件内容
    for i in range(len(datas)):
        # 获取标准答案
        gold = datas[i].get("gold", "")[0].replace(" ", "")# datas[i].get("gold")[0]
        # 获取预测答案
        pre = datas[i].get("answer", "").replace("<抽取结果>\n","")
        # 标记
        flag = 0
        # 存疑 判断分子类型
        if "Branch" in gold or "Ring" in gold:
            try:
                # 将gold以smiles格式解码
                gold1 = sf.decoder(gold)
                # 如果是smiles会转为"" ,不报错
                if gold1 != "":
                    gold = gold1
                    flag = 1
            except:
                gold = gold
        if flag == 1:
            try:
                # 将pre以smiles格式解码
                pre = sf.decoder(pre)
            except:
                pre = None

        # 将预测答案和标准答案添加到列表中
        pres.append([pre])
        golds.append([gold])

    # 定义评估指标
    metrics = ["exact_match"]
    # 计算评估指标
    # print(1111111111111,pres, golds)
    metric_results = calculate_smiles_metrics(pres, golds, metrics)
    # print(metric_results, type(metric_results))
    
    return {"分子设计": 100*(metric_results['num_t1_exact_match'])/metric_results['num_all']}

def calculate_f1_score(true_entities, predicted_entities):
    true_entities = true_entities.replace(" ", "")
    true_entities = set(true_entities.split(';'))
    predicted_entities = predicted_entities.replace(" ", "")
    predicted_entities = set(predicted_entities.split(';'))
    true_positive = len(true_entities & predicted_entities)
    # print(true_entities & predicted_entities)
    precision = true_positive / len(predicted_entities) if len(predicted_entities) > 0 else 0
    recall = true_positive / len(true_entities) if len(true_entities) > 0 else 0

    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    return f1_score

def entity_calculate(datas):
    pres = []
    golds = []
    for i in range(len(datas)):
        gold = datas[i].get("gold", "")
        pre = datas[i].get("answer", "").replace("<answer_start>", "").replace("<answer_end>", "")
        pres.append(pre)
        golds.append(gold)
    correct_count = 0
    total_count = len(golds)
    for i in range(len(golds)):
        true_output = golds[i].lower()
        if isinstance(pres[i], list):
            my_outputs = [item[:].lower() for item in pres[i] if isinstance(item, str)]  # 处理每个字符串元素
        else:
            my_output = pres[i][:].lower()
        print(111111111111,true_output)
        print(222222222222,my_output)
        f1_score = calculate_f1_score(true_output, my_output)
        correct_count += f1_score

    return {"信息抽取": 100 * correct_count / total_count}

def dict_to_excel(data_dict, file_path):
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    
    # 选择活动工作表
    sheet = workbook.active
    sheet.title = "Data"
    
    # 写入表头
    sheet.cell(row=1, column=1, value="Key")
    sheet.cell(row=1, column=2, value="Value")
    
    # 遍历字典，写入数据
    for i, (key, value) in enumerate(data_dict.items(), start=2):  # 从第二行开始
        sheet.cell(row=i, column=1, value=key)  # 将字典的键写入第一列
        sheet.cell(row=i, column=2, value=round(value, 2))  # 将字典的值写入第二列
    
    # 保存到指定路径
    workbook.save(file_path)

predict = "answer"
target = "target"

if __name__ == "__main__":
    ### You need to modify
    file_path = r"path/to/input/json"
    output_path = file_path+"_L1.xlsx"

    task_accuracy = {}
    
    datas = read_file(file_path)
    # print(datas)
    for k,v in datas.items():
        if k == "客观" and len(datas["客观"]) != 0:
            task_accuracy.update(objective_task(datas["客观"]))
        elif k == "主观" and len(datas["主观"]) != 0:
            task_accuracy.update(subjective_task(datas["主观"]))
        elif k == "分子翻译" and len(datas["分子翻译"]) != 0:
            task_accuracy.update(smiles2iupac(datas["分子翻译"]))
        elif k == "分子设计" and len(datas["分子设计"]) != 0:
            task_accuracy.update(molecule_design(datas["分子设计"]))
        elif k == "信息抽取" and len(datas["信息抽取"]) != 0:
            task_accuracy.update(entity_calculate(datas["信息抽取"]))
        else:
            task_accuracy.update(subjective_task(datas["主观"]))
            
    for task, acc in task_accuracy.items():
        # if acc == 0:
        print(f"{task}, {acc:.2f}%")
        print(f"Task: {task}, Accuracy: {acc:.2f}%")
    dict_to_excel(task_accuracy, output_path)